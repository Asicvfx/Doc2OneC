import csv
from io import BytesIO, TextIOWrapper

import fitz
from django.conf import settings
from openpyxl import load_workbook
from pypdf import PdfReader

from documents.models import Document

from .ocr import OCR_DISABLED_MESSAGE, extract_text_from_image_file, extract_text_from_pdf_page_image


PDF_NO_TEXT_MESSAGE = "PDF text extraction found no selectable text. OCR is required for scanned PDFs."
OCR_PLACEHOLDER = "OCR placeholder. Real OCR can be added here."


def parse_document_file(document: Document) -> str:
    document.file.open("rb")
    try:
        if document.file_type == Document.FileType.TXT:
            return _parse_text(document.file)
        if document.file_type == Document.FileType.CSV:
            return _parse_csv(document.file)
        if document.file_type == Document.FileType.XLSX:
            return _parse_xlsx(document.file)
        if document.file_type == Document.FileType.PDF:
            return _parse_pdf(document.file)
        if document.file_type == Document.FileType.IMAGE:
            return extract_text_from_image_file(document.file)
        return ""
    finally:
        document.file.close()


def _parse_text(file_obj) -> str:
    content = file_obj.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding).strip()
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace").strip()


def _parse_csv(file_obj) -> str:
    wrapper = TextIOWrapper(file_obj, encoding="utf-8-sig", newline="")
    sample = wrapper.read(2048)
    wrapper.seek(0)
    dialect = csv.Sniffer().sniff(sample, delimiters=",;") if sample else csv.excel
    reader = csv.DictReader(wrapper, dialect=dialect)
    rows = []
    for row in reader:
        rows.append("; ".join(f"{key}={value}" for key, value in row.items() if key and value))
    return "\n".join(rows).strip()


def _parse_xlsx(file_obj) -> str:
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed_rows = []
    for row in rows[1:]:
        pairs = []
        for header, value in zip(headers, row):
            if header and value is not None:
                pairs.append(f"{header}={value}")
        if pairs:
            parsed_rows.append("; ".join(pairs))
    return "\n".join(parsed_rows).strip()


def _parse_pdf(file_obj) -> str:
    pdf_bytes = file_obj.read()
    selectable_text = _parse_pdf_selectable_text(pdf_bytes)
    if selectable_text:
        return selectable_text

    ocr_text = _parse_pdf_with_ocr(pdf_bytes)
    if ocr_text and ocr_text != OCR_DISABLED_MESSAGE:
        return ocr_text
    return ocr_text or PDF_NO_TEXT_MESSAGE


def _parse_pdf_selectable_text(pdf_bytes: bytes) -> str:
    reader = PdfReader(BytesIO(pdf_bytes))
    pages = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        page_text = page_text.strip()
        if page_text:
            pages.append(page_text)
    return "\n\n".join(pages).strip()


def _parse_pdf_with_ocr(pdf_bytes: bytes) -> str:
    rendered_pages = _render_pdf_pages_to_png(pdf_bytes, max_pages=settings.OCR_MAX_PDF_PAGES)
    page_texts = []
    for image_bytes in rendered_pages:
        page_text = extract_text_from_pdf_page_image(image_bytes).strip()
        if page_text and page_text not in {OCR_DISABLED_MESSAGE, PDF_NO_TEXT_MESSAGE}:
            page_texts.append(page_text)
        elif page_text == OCR_DISABLED_MESSAGE:
            return OCR_DISABLED_MESSAGE
    return "\n\n".join(page_texts).strip() or PDF_NO_TEXT_MESSAGE


def _render_pdf_pages_to_png(pdf_bytes: bytes, max_pages: int) -> list[bytes]:
    rendered = []
    with fitz.open(stream=pdf_bytes, filetype="pdf") as pdf:
        for page in pdf[:max_pages]:
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            rendered.append(pixmap.tobytes("png"))
    return rendered
